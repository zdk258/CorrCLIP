"""Shared utilities: image unnormalize, xlsx result logger, and the BoundaryIoU metric."""
import os.path as osp
from collections import OrderedDict
from typing import Dict, List, Sequence

import cv2
import numpy as np
import openpyxl
import torch
from mmengine.logging import MMLogger, print_log
from PIL import Image
from prettytable import PrettyTable

from mmseg.evaluation.metrics.iou_metric import IoUMetric
from mmseg.registry import METRICS


class UnNormalize(object):
    def __init__(self, mean, std):
        self.mean = mean
        self.std = std

    def __call__(self, image):
        image2 = torch.clone(image)
        for t, m, s in zip(image2, self.mean, self.std):
            t.mul_(s).add_(m)
        return image2


def append_experiment_result(file_path, experiment_data):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    sheet = workbook.active

    if sheet['A1'].value is None:
        sheet['A1'] = 'Model'
        sheet['B1'] = 'CLIP'
        sheet['C1'] = 'DINO'
        sheet['D1'] = 'Dataset'
        sheet['E1'] = 'aAcc'
        sheet['F1'] = 'mIoU'
        sheet['G1'] = 'mAcc'

    last_row = sheet.max_row

    for index, result in enumerate(experiment_data, start=1):
        sheet.cell(row=last_row + index, column=1, value=result['Model'])
        sheet.cell(row=last_row + index, column=2, value=result['CLIP'])
        sheet.cell(row=last_row + index, column=3, value=result['DINO'])
        sheet.cell(row=last_row + index, column=4, value=result['Dataset'])
        sheet.cell(row=last_row + index, column=5, value=result['aAcc'])
        sheet.cell(row=last_row + index, column=6, value=result['mIoU'])
        sheet.cell(row=last_row + index, column=7, value=result['mAcc'])

    workbook.save(file_path)


# ===================== Boundary IoU evaluation metric =====================
# Mirrors LPOSS's `metrics/boundary_iou.py` (LPOSS, CVPR 2025; Cheng et al., CVPR 2021):
# 1. For each image, iterate over every class that appears in either pred or gt.
# 2. Build the per-class binary masks (pred==c, gt==c), with ignore pixels zeroed.
# 3. Apply boundary extraction (`mask - erode(mask)`, 3x3 kernel,
#    dilation = round(0.02 * image_diagonal)) to each binary mask.
# 4. Accumulate per-class boundary intersection (`(gt_b * dt_b) > 0`) and
#    union (`(gt_b + dt_b) > 0`) across the dataset.
# 5. Final per-class BoundaryIoU = sum_intersect_c / sum_union_c,
#    mBoundaryIoU = mean over classes (NaN-safe).
# It does NOT match the detectron2 SemSegEvaluator variant (which suffers from
# label-difference attribution and is typically 15-80 pts lower on multi-class datasets).
#
# Usage in a config:
#     test_evaluator = dict(
#         type='BoundaryIoUMetric',
#         iou_metrics=['mIoU', 'mBoundaryIoU'],
#         boundary_dilation_ratio=0.02,
#     )


def mask_to_boundary(mask: np.ndarray, dilation_ratio: float = 0.02) -> np.ndarray:
    """`mask - erode(mask)` with 3x3 min filter, padded by 0. Matches LPOSS / Cheng et al."""
    h, w = mask.shape
    img_diag = float(np.sqrt(h * h + w * w))
    dilation = max(1, int(round(dilation_ratio * img_diag)))
    kernel = np.ones((3, 3), dtype=np.uint8)
    padded = cv2.copyMakeBorder(mask, 1, 1, 1, 1, cv2.BORDER_CONSTANT, value=0)
    eroded = cv2.erode(padded, kernel, iterations=dilation)[1:h + 1, 1:w + 1]
    return mask - eroded


@METRICS.register_module()
class BoundaryIoUMetric(IoUMetric):
    """IoUMetric + per-class Boundary IoU (LPOSS / Cheng et al. style)."""

    BOUNDARY_KEY = 'mBoundaryIoU'

    def __init__(self,
                 iou_metrics: List[str] = ['mIoU', 'mBoundaryIoU'],
                 boundary_dilation_ratio: float = 0.02,
                 **kwargs) -> None:
        pixel_metrics = [m for m in iou_metrics if m != self.BOUNDARY_KEY]
        super().__init__(iou_metrics=pixel_metrics or ['mIoU'], **kwargs)
        self.need_boundary = self.BOUNDARY_KEY in iou_metrics
        self.boundary_dilation_ratio = boundary_dilation_ratio

    def process(self, data_batch: dict, data_samples: Sequence[dict]) -> None:
        num_classes = len(self.dataset_meta['classes'])
        for data_sample in data_samples:
            pred_label = data_sample['pred_sem_seg']['data'].squeeze()
            if not self.format_only:
                label = data_sample['gt_sem_seg']['data'].squeeze().to(pred_label)
                pixel_stats = self.intersect_and_union(
                    pred_label, label, num_classes, self.ignore_index)
                if self.need_boundary:
                    b_inter, b_union = self._boundary_intersect_and_union(
                        pred_label, label, num_classes, self.ignore_index)
                    self.results.append(pixel_stats + (b_inter, b_union))
                else:
                    self.results.append(pixel_stats)

            if self.output_dir is not None:
                basename = osp.splitext(osp.basename(data_sample['img_path']))[0]
                png_filename = osp.abspath(
                    osp.join(self.output_dir, f'{basename}.png'))
                output_mask = pred_label.cpu().numpy()
                if data_sample.get('reduce_zero_label', False):
                    output_mask = output_mask + 1
                Image.fromarray(output_mask.astype(np.uint8)).save(png_filename)

    def _boundary_intersect_and_union(self, pred_label: torch.Tensor,
                                      label: torch.Tensor, num_classes: int,
                                      ignore_index: int):
        """Per-class binary boundary IoU stats, following LPOSS's `boundary_iou`."""
        pred_np = pred_label.cpu().numpy().astype(np.int64).copy()
        label_np = label.cpu().numpy().astype(np.int64).copy()
        # mirror LPOSS: pred is forced to ignore wherever gt is ignore
        pred_np[label_np == ignore_index] = ignore_index

        b_intersect = torch.zeros(num_classes, dtype=torch.float64)
        b_union = torch.zeros(num_classes, dtype=torch.float64)

        present = (set(np.unique(label_np).tolist())
                   | set(np.unique(pred_np).tolist())) - {ignore_index}
        for c in present:
            if c < 0 or c >= num_classes:
                continue
            gt_bin = (label_np == c).astype(np.uint8)
            pred_bin = (pred_np == c).astype(np.uint8)
            gt_b = mask_to_boundary(gt_bin, self.boundary_dilation_ratio)
            pred_b = mask_to_boundary(pred_bin, self.boundary_dilation_ratio)
            inter = int(((gt_b * pred_b) > 0).sum())
            uni = int(((gt_b + pred_b) > 0).sum())
            b_intersect[c] = float(inter)
            b_union[c] = float(uni)
        return b_intersect, b_union

    def compute_metrics(self, results: list) -> Dict[str, float]:
        logger: MMLogger = MMLogger.get_current_instance()
        if self.format_only:
            logger.info(f'results are saved to {osp.dirname(self.output_dir)}')
            return OrderedDict()

        if self.need_boundary:
            packed = list(zip(*results))
            assert len(packed) == 6, f'expected 6-tuple per sample, got {len(packed)}'
            pixel_tuple = packed[:4]
            b_inter_list = packed[4]
            b_union_list = packed[5]
        else:
            pixel_tuple = tuple(zip(*results))
            b_inter_list = b_union_list = None

        total_area_intersect = sum(pixel_tuple[0])
        total_area_union = sum(pixel_tuple[1])
        total_area_pred_label = sum(pixel_tuple[2])
        total_area_label = sum(pixel_tuple[3])

        ret_metrics = self.total_area_to_metrics(
            total_area_intersect, total_area_union, total_area_pred_label,
            total_area_label, self.metrics, self.nan_to_num, self.beta)

        if self.need_boundary:
            total_b_inter = sum(b_inter_list).numpy()
            total_b_union = sum(b_union_list).numpy()
            biou = np.full_like(total_b_union, np.nan, dtype=np.float64)
            valid = total_b_union > 0
            biou[valid] = total_b_inter[valid] / total_b_union[valid]
            if self.nan_to_num is not None:
                biou = np.nan_to_num(biou, nan=self.nan_to_num)
            ret_metrics['BoundaryIoU'] = biou

        class_names = self.dataset_meta['classes']

        ret_metrics_summary = OrderedDict({
            ret_metric: np.round(np.nanmean(ret_metric_value) * 100, 2)
            for ret_metric, ret_metric_value in ret_metrics.items()
        })
        metrics_out: Dict[str, float] = dict()
        for key, val in ret_metrics_summary.items():
            if key == 'aAcc':
                metrics_out[key] = val
            else:
                metrics_out['m' + key] = val

        ret_metrics.pop('aAcc', None)
        ret_metrics_class = OrderedDict({
            ret_metric: np.round(ret_metric_value * 100, 2)
            for ret_metric, ret_metric_value in ret_metrics.items()
        })
        ret_metrics_class.update({'Class': class_names})
        ret_metrics_class.move_to_end('Class', last=False)
        table = PrettyTable()
        for key, val in ret_metrics_class.items():
            table.add_column(key, val)

        print_log('per class results:', logger)
        print_log('\n' + table.get_string(), logger=logger)

        return metrics_out
